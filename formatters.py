"""
Excel formatting utilities
"""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from config import Config
import logging

logger = logging.getLogger(__name__)

class ExcelFormatter:
    """Excel formatting utilities"""

    @staticmethod
    def format_header(worksheet, config=Config):
        """
        Format header row with styling

        Args:
            worksheet: openpyxl worksheet object
            config: Configuration object
        """
        header_font = Font(
            name=config.HEADER_FONT,
            size=config.HEADER_FONT_SIZE,
            bold=True,
            color="FFFFFF"
        )

        header_fill = PatternFill(
            start_color=config.HEADER_COLOR,
            end_color=config.HEADER_COLOR,
            fill_type="solid"
        )

        header_border = Border(
            bottom=Side(style='thick', color=config.HEADER_COLOR)
        )

        header_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        # Apply to first row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment

        # Set row height
        worksheet.row_dimensions[1].height = config.ROW_HEIGHT_HEADER

        logger.info("Header formatting applied")

    @staticmethod
    def format_currency_column(worksheet, column_letter, start_row=2, config=Config):
        """
        Format a column as currency

        Args:
            worksheet: openpyxl worksheet object
            column_letter: Column letter (e.g., 'D')
            start_row: Starting row number
            config: Configuration object
        """
        max_row = worksheet.max_row

        for row in range(start_row, max_row + 1):
            cell = worksheet[f"{column_letter}{row}"]
            cell.number_format = config.CURRENCY_FORMAT

        logger.info(f"Currency formatting applied to column {column_letter}")

    @staticmethod
    def format_percentage_column(worksheet, column_letter, start_row=2, config=Config):
        """
        Format a column as percentage

        Args:
            worksheet: openpyxl worksheet object
            column_letter: Column letter (e.g., 'E')
            start_row: Starting row number
            config: Configuration object
        """
        max_row = worksheet.max_row

        for row in range(start_row, max_row + 1):
            cell = worksheet[f"{column_letter}{row}"]
            cell.number_format = config.PERCENTAGE_FORMAT

        logger.info(f"Percentage formatting applied to column {column_letter}")

    @staticmethod
    def format_date_column(worksheet, column_letter, start_row=2, config=Config):
        """
        Format a column as date

        Args:
            worksheet: openpyxl worksheet object
            column_letter: Column letter (e.g., 'A')
            start_row: Starting row number
            config: Configuration object
        """
        max_row = worksheet.max_row

        for row in range(start_row, max_row + 1):
            cell = worksheet[f"{column_letter}{row}"]
            cell.number_format = config.DATE_FORMAT

        logger.info(f"Date formatting applied to column {column_letter}")

    @staticmethod
    def auto_adjust_column_width(worksheet, min_width=10, max_width=50):
        """
        Auto-adjust column widths based on content

        Args:
            worksheet: openpyxl worksheet object
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        logger.info("Column widths auto-adjusted")

    @staticmethod
    def apply_alternating_rows(worksheet, start_row=2, color="F2F2F2"):
        """
        Apply alternating row colors for readability

        Args:
            worksheet: openpyxl worksheet object
            start_row: Starting row number
            color: Fill color for alternate rows
        """
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for row in range(start_row, max_row + 1):
            if row % 2 == 0:
                for col in range(1, max_col + 1):
                    worksheet.cell(row=row, column=col).fill = fill

        logger.info("Alternating row colors applied")

    @staticmethod
    def apply_borders(worksheet, start_row=1):
        """
        Apply borders to all cells

        Args:
            worksheet: openpyxl worksheet object
            start_row: Starting row number
        """
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for row in range(start_row, max_row + 1):
            for col in range(1, max_col + 1):
                worksheet.cell(row=row, column=col).border = thin_border

        logger.info("Borders applied to all cells")

    @staticmethod
    def add_title(worksheet, title, config=Config):
        """
        Add a title row at the top

        Args:
            worksheet: openpyxl worksheet object
            title: Title text
            config: Configuration object
        """
        # Insert a new row at the top
        worksheet.insert_rows(1)

        # Merge cells for title
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)

        # Style the title
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(
            name=config.HEADER_FONT,
            size=config.HEADER_FONT_SIZE + 4,
            bold=True,
            color=config.HEADER_COLOR
        )
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set row height
        worksheet.row_dimensions[1].height = 40

        logger.info(f"Title added: {title}")

    @staticmethod
    def freeze_panes(worksheet, row=2, column=1):
        """
        Freeze panes for easier scrolling

        Args:
            worksheet: openpyxl worksheet object
            row: Row to freeze after
            column: Column to freeze after
        """
        cell = f"{get_column_letter(column)}{row}"
        worksheet.freeze_panes = cell
        logger.info(f"Panes frozen at {cell}")

    @staticmethod
    def add_autofilter(worksheet, start_row=1):
        """
        Add autofilter to headers

        Args:
            worksheet: openpyxl worksheet object
            start_row: Header row number
        """
        max_col = worksheet.max_column
        max_row = worksheet.max_row

        filter_range = f"A{start_row}:{get_column_letter(max_col)}{max_row}"
        worksheet.auto_filter.ref = filter_range

        logger.info(f"Autofilter applied: {filter_range}")
