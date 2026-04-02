"""
Main Excel Report Generator
Creates professional Excel reports from data
"""
import argparse
import logging
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from config import Config
from data_loader import DataLoader
from formatters import ExcelFormatter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('report_generator.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Main report generation class"""

    def __init__(self, config=Config):
        self.config = config
        self.loader = DataLoader()
        self.formatter = ExcelFormatter()

    def generate_basic_report(self, data, output_file, title=None):
        """
        Generate a basic formatted Excel report

        Args:
            data: pandas DataFrame
            output_file: Output file path
            title: Report title (optional)
        """
        logger.info(f"Generating basic report: {output_file}")

        # Write to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name='Data', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Data']

            # Apply formatting
            self.formatter.format_header(worksheet, self.config)
            self.formatter.auto_adjust_column_width(worksheet)
            self.formatter.apply_alternating_rows(worksheet)
            self.formatter.apply_borders(worksheet)

            if self.config.AUTO_FILTER:
                self.formatter.add_autofilter(worksheet)

            if self.config.FREEZE_PANES:
                self.formatter.freeze_panes(worksheet)

            if title:
                self.formatter.add_title(worksheet, title, self.config)

        logger.info(f"Report generated successfully: {output_file}")

    def generate_sales_report(self, data, output_file):
        """
        Generate a sales report with charts and summary

        Args:
            data: pandas DataFrame with sales data
            output_file: Output file path
        """
        logger.info(f"Generating sales report: {output_file}")

        # Validate required columns
        required_cols = ['Date', 'Region', 'Product', 'Sales', 'Profit']
        self.loader.validate_data(data, required_cols)

        # Create summary statistics
        summary = {
            'Metric': ['Total Sales', 'Total Profit', 'Avg Daily Sales', 'Avg Profit Margin %', 'Total Transactions'],
            'Value': [
                data['Sales'].sum(),
                data['Profit'].sum(),
                data['Sales'].mean(),
                (data['Profit'].sum() / data['Sales'].sum() * 100),
                len(data)
            ]
        }
        summary_df = pd.DataFrame(summary)

        # Sales by region
        region_summary = data.groupby('Region').agg({
            'Sales': 'sum',
            'Profit': 'sum',
            'Quantity': 'sum'
        }).reset_index()

        # Sales by product
        product_summary = data.groupby('Product').agg({
            'Sales': 'sum',
            'Profit': 'sum',
            'Quantity': 'sum'
        }).reset_index()

        # Write to Excel with multiple sheets
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            self._format_summary_sheet(writer.sheets['Summary'])

            # Detailed data
            data.to_excel(writer, sheet_name='Detailed Data', index=False)
            self._format_data_sheet(writer.sheets['Detailed Data'], data)

            # Region analysis
            region_summary.to_excel(writer, sheet_name='By Region', index=False)
            self._format_data_sheet(writer.sheets['By Region'], region_summary)
            self._add_bar_chart(writer.sheets['By Region'], 'Sales by Region')

            # Product analysis
            product_summary.to_excel(writer, sheet_name='By Product', index=False)
            self._format_data_sheet(writer.sheets['By Product'], product_summary)
            self._add_bar_chart(writer.sheets['By Product'], 'Sales by Product')

        logger.info(f"Sales report generated successfully: {output_file}")

    def _format_summary_sheet(self, worksheet):
        """Format the summary sheet"""
        self.formatter.format_header(worksheet, self.config)
        self.formatter.format_currency_column(worksheet, 'B', start_row=2)
        self.formatter.auto_adjust_column_width(worksheet)
        self.formatter.apply_borders(worksheet)

    def _format_data_sheet(self, worksheet, data):
        """Format a data sheet with appropriate column formats"""
        self.formatter.format_header(worksheet, self.config)

        # Identify currency columns
        currency_cols = ['Sales', 'Profit', 'Cost']
        percentage_cols = ['Profit_Margin']
        date_cols = ['Date']

        for idx, col in enumerate(data.columns, start=1):
            col_letter = self.formatter.__class__.__dict__['format_currency_column'].__code__.co_varnames[1]
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(idx)

            if col in currency_cols:
                self.formatter.format_currency_column(worksheet, col_letter)
            elif col in percentage_cols:
                self.formatter.format_percentage_column(worksheet, col_letter)
            elif col in date_cols:
                self.formatter.format_date_column(worksheet, col_letter)

        self.formatter.auto_adjust_column_width(worksheet)
        self.formatter.apply_alternating_rows(worksheet)
        self.formatter.apply_borders(worksheet)
        self.formatter.freeze_panes(worksheet)
        self.formatter.add_autofilter(worksheet)

    def _add_bar_chart(self, worksheet, title):
        """Add a bar chart to the worksheet"""
        chart = BarChart()
        chart.title = title
        chart.style = self.config.CHART_STYLE
        chart.width = self.config.CHART_WIDTH
        chart.height = self.config.CHART_HEIGHT

        # Data references
        data = Reference(worksheet, min_col=2, min_row=1, max_row=worksheet.max_row, max_col=2)
        categories = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Place chart
        worksheet.add_chart(chart, "E2")

        logger.info(f"Chart added: {title}")

    def generate_from_csv(self, csv_file, output_file, report_type='basic'):
        """
        Generate report from CSV file

        Args:
            csv_file: Input CSV file path
            output_file: Output Excel file path
            report_type: Type of report ('basic' or 'sales')
        """
        # Load data
        data = self.loader.load_csv(csv_file)

        # Generate appropriate report
        if report_type == 'sales':
            self.generate_sales_report(data, output_file)
        else:
            title = self.config.get_report_title("Data Report")
            self.generate_basic_report(data, output_file, title)

    def generate_sample_report(self, output_file):
        """
        Generate a sample report using generated data

        Args:
            output_file: Output Excel file path
        """
        logger.info("Generating sample sales report with generated data")

        # Create sample data
        data = self.loader.create_sample_data()

        # Generate sales report
        self.generate_sales_report(data, output_file)


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description='Excel Report Generator')
    parser.add_argument('--input', type=str, help='Input CSV file path')
    parser.add_argument('--output', type=str, default='report.xlsx', help='Output Excel file path')
    parser.add_argument('--template', type=str, choices=['basic', 'sales'], default='basic',
                       help='Report template type')
    parser.add_argument('--sample', action='store_true', help='Generate sample report')

    args = parser.parse_args()

    # Create generator
    generator = ReportGenerator()

    try:
        if args.sample:
            # Generate sample report
            logger.info("Generating sample report")
            generator.generate_sample_report(args.output)
        elif args.input:
            # Generate from input file
            generator.generate_from_csv(args.input, args.output, args.template)
        else:
            logger.error("Please provide --input file or use --sample flag")
            parser.print_help()
            return

        logger.info(f"\n{'='*60}")
        logger.info(f"SUCCESS: Report generated at {args.output}")
        logger.info(f"{'='*60}\n")

    except Exception as e:
        logger.error(f"Report generation failed: {str(e)}", exc_info=True)
        raise


if __name__ == '__main__':
    main()
